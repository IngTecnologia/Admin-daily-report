"""
Manejador de archivos Excel para el sistema de reportes diarios
Implementa la estructura de BD especificada en el README
"""
import os
import uuid
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import pytz

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from .config import settings, EXCEL_SCHEMA
from .models import DailyReportCreate, DailyReportResponse, IncidentResponse, MovementResponse

# Timezone de Bogotá (GMT-5)
BOGOTA_TZ = pytz.timezone('America/Bogota')

def get_bogota_now() -> datetime:
    """Obtiene la fecha y hora actual en timezone de Bogotá (sin timezone para Excel)"""
    # Excel no soporta timezones, convertir a naive datetime
    return datetime.now(BOGOTA_TZ).replace(tzinfo=None)


class ExcelHandler:
    """Manejador principal para operaciones con Excel"""
    
    def __init__(self, file_path: Optional[Path] = None):
        self.file_path = file_path or settings.excel_file_path
        self.sheets = settings.excel_sheets
        self._ensure_file_exists()
        
    def _ensure_file_exists(self) -> None:
        """Crear el archivo Excel y sus hojas si no existe"""
        if not self.file_path.exists():
            self._create_initial_file()
        else:
            self._validate_structure()
    
    def _create_initial_file(self) -> None:
        """Crear archivo Excel inicial con la estructura correcta"""
        workbook = Workbook()
        
        # Remover la hoja por defecto
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Crear cada hoja segun el esquema
        for sheet_key, sheet_name in self.sheets.items():
            if sheet_key in EXCEL_SCHEMA:
                worksheet = workbook.create_sheet(sheet_name)
                self._setup_sheet_headers(worksheet, sheet_key)
        
        # Crear hoja de configuracion inicial
        self._create_config_sheet(workbook)
        
        # Guardar archivo
        workbook.save(self.file_path)
        print(f"Archivo Excel creado: {self.file_path}")
    
    def _setup_sheet_headers(self, worksheet, sheet_key: str) -> None:
        """Configurar encabezados y estilos de una hoja"""
        schema = EXCEL_SCHEMA[sheet_key]
        headers = schema["columns"]
        
        # Establecer encabezados
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            
            # Estilo para encabezados
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar anchos de columna
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_config_sheet(self, workbook: Workbook) -> None:
        """Crear hoja de configuracion inicial"""
        config_sheet = workbook.create_sheet(self.sheets["configuracion"])
        
        # Configurar headers
        headers = ["Clave", "Valor", "Descripcion", "Fecha_Modificacion"]
        for col_num, header in enumerate(headers, 1):
            config_sheet.cell(row=1, column=col_num, value=header)
        
        # Configuraciones iniciales
        config_data = [
            ["version_sistema", "1.0.0", "Version del sistema", get_bogota_now()],
            ["max_reportes_dia", "10", "Maximo reportes por dia por admin", get_bogota_now()],
            ["fecha_creacion", get_bogota_now().isoformat(), "Fecha de creacion del sistema", get_bogota_now()],
            ["backup_enabled", "true", "Backups automaticos habilitados", get_bogota_now()],
        ]
        
        # Agregar datos de configuracion
        for row_num, data in enumerate(config_data, 2):
            for col_num, value in enumerate(data, 1):
                config_sheet.cell(row=row_num, column=col_num, value=value)
    
    def _validate_structure(self) -> None:
        """Validar que el archivo Excel tenga la estructura correcta"""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            required_sheets = set(self.sheets.values())
            existing_sheets = set(workbook.sheetnames)
            
            # Verificar que todas las hojas requeridas existan
            missing_sheets = required_sheets - existing_sheets
            if missing_sheets:
                print(f"Advertencia: Faltan hojas en Excel: {missing_sheets}")
                # Agregar hojas faltantes
                for sheet_name in missing_sheets:
                    sheet_key = next(k for k, v in self.sheets.items() if v == sheet_name)
                    worksheet = workbook.create_sheet(sheet_name)
                    self._setup_sheet_headers(worksheet, sheet_key)
                
                workbook.save(self.file_path)
                
        except Exception as e:
            print(f"Error validando estructura Excel: {e}")
            # Si hay problemas, recrear el archivo
            self._create_initial_file()
    
    def generate_report_id(self) -> str:
        """Generar ID unico para reporte segun formato especificado"""
        today = get_bogota_now()
        date_str = today.strftime("%Y%m%d")
        
        # Contar reportes del dia
        existing_reports = self.get_reports_by_date(today.date())
        count = len(existing_reports) + 1
        
        return f"RPT-{date_str}-{count:03d}"
    
    def save_report(self, report: DailyReportCreate, client_info: Dict[str, str]) -> DailyReportResponse:
        """Guardar reporte completo en Excel"""
        try:
            report_id = self.generate_report_id()
            timestamp = get_bogota_now()
            
            # Guardar reporte principal
            self._save_main_report(report_id, report, timestamp, client_info)
            
            # Guardar incidencias
            incident_responses = []
            if report.incidencias:
                incident_responses = self._save_incidents(report_id, report.incidencias, timestamp)
            
            # Guardar movimientos
            movement_responses = []
            if report.ingresos_retiros:
                movement_responses = self._save_movements(report_id, report.ingresos_retiros, timestamp)
            
            # Crear respuesta
            response = DailyReportResponse(
                id=report_id,
                fecha_creacion=timestamp,
                administrador=report.administrador.value,
                cliente_operacion=report.cliente_operacion.value,
                horas_diarias=report.horas_diarias,
                personal_staff=report.personal_staff,
                personal_base=report.personal_base,
                cantidad_incidencias=len(report.incidencias),
                cantidad_ingresos_retiros=len(report.ingresos_retiros),
                hechos_relevantes=report.hechos_relevantes or "",
                estado="Completado",
                incidencias=incident_responses,
                ingresos_retiros=movement_responses
            )
            
            return response
            
        except Exception as e:
            print(f"Error guardando reporte: {e}")
            raise Exception(f"Error al guardar el reporte: {str(e)}")
    
    def _save_main_report(self, report_id: str, report: DailyReportCreate, 
                         timestamp: datetime, client_info: Dict[str, str]) -> None:
        """Guardar reporte principal en hoja Reportes"""
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook[self.sheets["reportes"]]
        
        # Encontrar siguiente fila vacia
        next_row = sheet.max_row + 1
        
        # Datos del reporte segun esquema
        row_data = [
            report_id,
            timestamp,
            report.administrador.value,
            report.cliente_operacion.value,
            report.horas_diarias,
            report.personal_staff,
            report.personal_base,
            len(report.incidencias),
            len(report.ingresos_retiros),
            report.hechos_relevantes or "",
            "Completado",
            client_info.get("ip", "Unknown"),
            client_info.get("user_agent", "Unknown")
        ]
        
        # Escribir datos
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=next_row, column=col_num, value=value)
        
        workbook.save(self.file_path)
    
    def _save_incidents(self, report_id: str, incidents: List, timestamp: datetime) -> List[IncidentResponse]:
        """Guardar incidencias en hoja Incidencias"""
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook[self.sheets["incidencias"]]
        
        responses = []
        
        for incident_num, incident in enumerate(incidents, 1):
            next_row = sheet.max_row + 1
            
            # Datos de la incidencia segun esquema
            row_data = [
                report_id,
                incident_num,
                incident.tipo.value,
                incident.nombre_empleado,
                incident.fecha_fin,
                timestamp
            ]
            
            # Escribir datos
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=next_row, column=col_num, value=value)
            
            # Crear respuesta
            response = IncidentResponse(
                id=incident_num,
                tipo=incident.tipo,
                nombre_empleado=incident.nombre_empleado,
                fecha_fin=incident.fecha_fin,
                fecha_registro=timestamp
            )
            responses.append(response)
        
        workbook.save(self.file_path)
        return responses
    
    def _save_movements(self, report_id: str, movements: List, timestamp: datetime) -> List[MovementResponse]:
        """Guardar movimientos en hoja Ingresos_Retiros"""
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook[self.sheets["ingresos_retiros"]]
        
        responses = []
        
        for movement_num, movement in enumerate(movements, 1):
            next_row = sheet.max_row + 1
            
            # Datos del movimiento segun esquema
            row_data = [
                report_id,
                movement_num,
                movement.nombre_empleado,
                movement.cargo,
                movement.estado.value,
                timestamp
            ]
            
            # Escribir datos
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=next_row, column=col_num, value=value)
            
            # Crear respuesta
            response = MovementResponse(
                id=movement_num,
                nombre_empleado=movement.nombre_empleado,
                cargo=movement.cargo,
                estado=movement.estado,
                fecha_registro=timestamp
            )
            responses.append(response)
        
        workbook.save(self.file_path)
        return responses
    
    def get_reports_by_date(self, target_date: date) -> List[Dict[str, Any]]:
        """Obtener reportes por fecha"""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook[self.sheets["reportes"]]
            
            reports = []
            headers = [cell.value for cell in sheet[1]]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:  # Fila vacia
                    break
                    
                row_dict = dict(zip(headers, row))
                
                # Verificar fecha
                fecha_creacion = row_dict.get('Fecha_Creacion')
                if isinstance(fecha_creacion, datetime):
                    # Convertir a timezone de Bogotá si no tiene timezone
                    if fecha_creacion.tzinfo is None:
                        fecha_creacion = BOGOTA_TZ.localize(fecha_creacion)
                    fecha_creacion = fecha_creacion.astimezone(BOGOTA_TZ)
                    if fecha_creacion.date() == target_date:
                        reports.append(row_dict)
                elif isinstance(fecha_creacion, str):
                    try:
                        parsed_date = datetime.fromisoformat(fecha_creacion).date()
                        if parsed_date == target_date:
                            reports.append(row_dict)
                    except:
                        continue
            
            return reports
            
        except Exception as e:
            print(f"Error obteniendo reportes por fecha: {e}")
            return []
    
    def get_all_reports(self, filters: Optional[Dict] = None) -> List[Dict[str, Any]]:
        """Obtener todos los reportes con filtros opcionales"""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook[self.sheets["reportes"]]
            
            reports = []
            headers = [cell.value for cell in sheet[1]]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:  # Fila vacia
                    break
                    
                row_dict = dict(zip(headers, row))
                
                # Aplicar filtros si existen
                if filters:
                    if not self._apply_filters(row_dict, filters):
                        continue
                
                reports.append(row_dict)
            
            return reports
            
        except Exception as e:
            print(f"Error obteniendo reportes: {e}")
            return []
    
    def _apply_filters(self, report: Dict, filters: Dict) -> bool:
        """Aplicar filtros a un reporte"""
        # Filtro por administrador
        if filters.get('administrador') and report.get('Administrador') != filters['administrador']:
            return False
        
        # Filtro por cliente/operacion
        if filters.get('cliente') and report.get('Cliente_Operacion') != filters['cliente']:
            return False
        
        # Filtros de fecha
        if filters.get('fecha_inicio') or filters.get('fecha_fin'):
            try:
                # Obtener fecha del reporte (solo fecha, sin hora)
                report_datetime = report.get('Fecha_Creacion')
                if isinstance(report_datetime, str):
                    report_datetime = datetime.fromisoformat(report_datetime.replace('Z', '+00:00'))
                elif not isinstance(report_datetime, datetime):
                    return False
                
                report_date = report_datetime.date()
                
                # Filtro fecha inicio
                if filters.get('fecha_inicio'):
                    fecha_inicio = filters['fecha_inicio']
                    if isinstance(fecha_inicio, str):
                        fecha_inicio = datetime.fromisoformat(fecha_inicio).date()
                    if report_date < fecha_inicio:
                        return False
                
                # Filtro fecha fin
                if filters.get('fecha_fin'):
                    fecha_fin = filters['fecha_fin']
                    if isinstance(fecha_fin, str):
                        fecha_fin = datetime.fromisoformat(fecha_fin).date()
                    if report_date > fecha_fin:
                        return False
                        
            except (ValueError, TypeError):
                # Si hay error procesando fechas, incluir el reporte
                pass
            
        return True
    
    def get_report_incidents(self, report_id: str) -> List[Dict[str, Any]]:
        """Obtener incidencias de un reporte específico"""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook[self.sheets["incidencias"]]
            
            incidents = []
            headers = [cell.value for cell in sheet[1]]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:  # Fila vacía
                    break
                    
                row_dict = dict(zip(headers, row))
                
                # Filtrar solo las incidencias de este reporte
                if row_dict.get('ID_Reporte') == report_id:
                    incidents.append({
                        'tipo': row_dict.get('Tipo_Incidencia'),
                        'nombre_empleado': row_dict.get('Nombre_Empleado'),
                        'fecha_fin': row_dict.get('Fecha_Fin_Novedad'),
                        'fecha_registro': row_dict.get('Fecha_Registro')
                    })
            
            return incidents
            
        except Exception as e:
            print(f"Error obteniendo incidencias del reporte {report_id}: {e}")
            return []
    
    def get_report_movements(self, report_id: str) -> List[Dict[str, Any]]:
        """Obtener movimientos de personal de un reporte específico"""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook[self.sheets["ingresos_retiros"]]
            
            movements = []
            headers = [cell.value for cell in sheet[1]]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:  # Fila vacía
                    break
                    
                row_dict = dict(zip(headers, row))
                
                # Filtrar solo los movimientos de este reporte
                if row_dict.get('ID_Reporte') == report_id:
                    movements.append({
                        'nombre_empleado': row_dict.get('Nombre_Empleado'),
                        'cargo': row_dict.get('Cargo'),
                        'estado': row_dict.get('Estado'),
                        'fecha_registro': row_dict.get('Fecha_Registro')
                    })
            
            return movements
            
        except Exception as e:
            print(f"Error obteniendo movimientos del reporte {report_id}: {e}")
            return []
    
    def get_analytics_data(self) -> Dict[str, Any]:
        """Obtener datos para analytics del dashboard"""
        try:
            today = get_bogota_now().date()
            all_reports = self.get_all_reports()
            reports_today = self.get_reports_by_date(today)
            
            # Calcular estadisticas basicas
            total_reportes = len(all_reports)
            reportes_hoy = len(reports_today)
            
            # Promedio de horas diarias
            total_horas = sum(r.get('Horas_Diarias', 0) for r in all_reports if r.get('Horas_Diarias'))
            promedio_horas = total_horas / total_reportes if total_reportes > 0 else 0
            
            # Total incidencias del mes
            current_month = today.month
            current_year = today.year
            incidencias_mes = 0
            
            for report in all_reports:
                fecha_creacion = report.get('Fecha_Creacion')
                if isinstance(fecha_creacion, datetime):
                    # Convertir a timezone de Bogotá si no tiene timezone
                    if fecha_creacion.tzinfo is None:
                        fecha_creacion = BOGOTA_TZ.localize(fecha_creacion)
                    fecha_creacion = fecha_creacion.astimezone(BOGOTA_TZ)
                    if fecha_creacion.month == current_month and fecha_creacion.year == current_year:
                        incidencias_mes += report.get('Cantidad_Incidencias', 0)
            
            # Administradores activos (unicos que han reportado)
            administradores_activos = len(set(r.get('Administrador') for r in all_reports if r.get('Administrador')))
            
            return {
                "total_reportes": total_reportes,
                "reportes_hoy": reportes_hoy,
                "promedio_horas_diarias": round(promedio_horas, 2),
                "total_incidencias_mes": incidencias_mes,
                "administradores_activos": administradores_activos,
                "graficos": {
                    # Placeholder para graficos que se pueden implementar despues
                    "reportes_por_dia": [],
                    "incidencias_por_tipo": [],
                    "personal_por_operacion": []
                }
            }
            
        except Exception as e:
            print(f"Error obteniendo analytics: {e}")
            return {
                "total_reportes": 0,
                "reportes_hoy": 0,
                "promedio_horas_diarias": 0,
                "total_incidencias_mes": 0,
                "administradores_activos": 0,
                "graficos": {}
            }
    
    def backup_file(self) -> bool:
        """Crear backup del archivo Excel"""
        try:
            backup_name = f"backup_{get_bogota_now().strftime('%Y%m%d_%H%M%S')}_{self.file_path.name}"
            backup_path = self.file_path.parent / "backups" / backup_name
            backup_path.parent.mkdir(exist_ok=True)
            
            import shutil
            shutil.copy2(self.file_path, backup_path)
            print(f"Backup creado: {backup_path}")
            return True
            
        except Exception as e:
            print(f"Error creando backup: {e}")
            return False
    
    def delete_report(self, report_id: str) -> bool:
        """
        Eliminar un reporte y sus registros relacionados
        
        Args:
            report_id: ID del reporte a eliminar
            
        Returns:
            bool: True si se eliminó correctamente, False en caso contrario
        """
        try:
            # Crear backup antes de eliminar
            if not self.backup_file():
                print("Advertencia: No se pudo crear backup antes de eliminar")
            
            workbook = openpyxl.load_workbook(self.file_path)
            
            # Eliminar de la hoja principal de reportes
            reportes_sheet = workbook[self.sheets["reportes"]]
            report_row_to_delete = None
            
            for row_idx, row in enumerate(reportes_sheet.iter_rows(min_row=2), start=2):
                if row[0].value == report_id:  # Columna A contiene el ID
                    report_row_to_delete = row_idx
                    break
            
            if report_row_to_delete is None:
                print(f"Reporte {report_id} no encontrado en la hoja de reportes")
                workbook.close()
                return False
            
            # Eliminar la fila del reporte
            reportes_sheet.delete_rows(report_row_to_delete, 1)
            print(f"Reporte {report_id} eliminado de la hoja de reportes")
            
            # Eliminar incidencias relacionadas
            if "incidencias" in self.sheets and self.sheets["incidencias"] in workbook.sheetnames:
                incidencias_sheet = workbook[self.sheets["incidencias"]]
                rows_to_delete = []
                
                for row_idx, row in enumerate(incidencias_sheet.iter_rows(min_row=2), start=2):
                    if row[0].value == report_id:  # Columna A contiene ID_Reporte
                        rows_to_delete.append(row_idx)
                
                # Eliminar filas en orden inverso para no afectar indices
                for row_idx in sorted(rows_to_delete, reverse=True):
                    incidencias_sheet.delete_rows(row_idx, 1)
                
                print(f"Eliminadas {len(rows_to_delete)} incidencias del reporte {report_id}")
            
            # Eliminar movimientos de personal relacionados
            if "ingresos_retiros" in self.sheets and self.sheets["ingresos_retiros"] in workbook.sheetnames:
                movimientos_sheet = workbook[self.sheets["ingresos_retiros"]]
                rows_to_delete = []
                
                for row_idx, row in enumerate(movimientos_sheet.iter_rows(min_row=2), start=2):
                    if row[0].value == report_id:  # Columna A contiene ID_Reporte
                        rows_to_delete.append(row_idx)
                
                # Eliminar filas en orden inverso para no afectar indices
                for row_idx in sorted(rows_to_delete, reverse=True):
                    movimientos_sheet.delete_rows(row_idx, 1)
                
                print(f"Eliminados {len(rows_to_delete)} movimientos del reporte {report_id}")
            
            # Guardar cambios
            workbook.save(self.file_path)
            workbook.close()
            
            print(f"Reporte {report_id} y todos sus registros relacionados eliminados exitosamente")
            return True
            
        except Exception as e:
            print(f"Error eliminando reporte {report_id}: {e}")
            try:
                workbook.close()
            except:
                pass
            return False

    def update_report(self, report_id: str, update_data: Dict[str, Any]) -> bool:
        """
        Actualizar un reporte existente
        
        Args:
            report_id: ID del reporte a actualizar
            update_data: Diccionario con los campos a actualizar y sus nuevos valores
            
        Returns:
            bool: True si se actualizó correctamente, False en caso contrario
        """
        try:
            # Crear backup antes de actualizar
            if not self.backup_file():
                print("Advertencia: No se pudo crear backup antes de actualizar")
            
            workbook = openpyxl.load_workbook(self.file_path)
            
            # Encontrar el reporte en la hoja principal
            reportes_sheet = workbook[self.sheets["reportes"]]
            report_row_to_update = None
            header_row = None
            
            # Obtener encabezados de la primera fila
            header_row = [cell.value for cell in reportes_sheet[1]]
            
            # Buscar la fila del reporte
            for row_idx, row in enumerate(reportes_sheet.iter_rows(min_row=2), start=2):
                if row[0].value == report_id:  # Columna A contiene el ID
                    report_row_to_update = row_idx
                    break
            
            if report_row_to_update is None:
                print(f"Reporte {report_id} no encontrado en la hoja de reportes")
                workbook.close()
                return False
            
            # Actualizar los campos especificados
            for field_name, new_value in update_data.items():
                # Encontrar el índice de la columna para este campo
                column_index = None
                for idx, header in enumerate(header_row):
                    if header == field_name:
                        column_index = idx + 1  # openpyxl usa índices base 1
                        break
                
                if column_index is not None:
                    reportes_sheet.cell(row=report_row_to_update, column=column_index).value = new_value
                    print(f"Actualizado {field_name} = {new_value} en reporte {report_id}")
                else:
                    print(f"Campo {field_name} no encontrado en los encabezados")
            
            # Guardar cambios
            workbook.save(self.file_path)
            workbook.close()
            
            print(f"Reporte {report_id} actualizado exitosamente")
            return True
            
        except Exception as e:
            print(f"Error actualizando reporte {report_id}: {e}")
            try:
                workbook.close()
            except:
                pass
            return False

    def update_report_incidents(self, report_id: str, incidents: List[Any]) -> bool:
        """
        Actualizar las incidencias de un reporte específico
        
        Args:
            report_id: ID del reporte
            incidents: Lista de incidencias actualizadas
            
        Returns:
            bool: True si se actualizó correctamente
        """
        try:
            # Crear backup antes de actualizar
            if not self.backup_file():
                print("Advertencia: No se pudo crear backup antes de actualizar incidencias")
            
            workbook = openpyxl.load_workbook(self.file_path)
            
            # Eliminar incidencias existentes del reporte
            if "incidencias" in self.sheets and self.sheets["incidencias"] in workbook.sheetnames:
                incidencias_sheet = workbook[self.sheets["incidencias"]]
                rows_to_delete = []
                
                for row_idx, row in enumerate(incidencias_sheet.iter_rows(min_row=2), start=2):
                    if row[0].value == report_id:  # Columna A contiene ID_Reporte
                        rows_to_delete.append(row_idx)
                
                # Eliminar filas en orden inverso para no afectar índices
                for row_idx in sorted(rows_to_delete, reverse=True):
                    incidencias_sheet.delete_rows(row_idx, 1)
                
                # Agregar las nuevas incidencias
                for incident in incidents:
                    # Manejar tanto objetos Pydantic como diccionarios
                    if hasattr(incident, 'dict'):
                        incident_dict = incident.dict()
                    else:
                        incident_dict = incident
                    
                    new_row = [
                        report_id,
                        incident_dict.get('tipo', ''),
                        incident_dict.get('nombre_empleado', ''),
                        incident_dict.get('fecha_fin', '')
                    ]
                    incidencias_sheet.append(new_row)
                
                print(f"Actualizadas {len(incidents)} incidencias para reporte {report_id}")
            
            # Guardar cambios
            workbook.save(self.file_path)
            workbook.close()
            
            return True
            
        except Exception as e:
            print(f"Error actualizando incidencias del reporte {report_id}: {e}")
            try:
                workbook.close()
            except:
                pass
            return False

    def update_report_movements(self, report_id: str, movements: List[Any]) -> bool:
        """
        Actualizar los movimientos de personal de un reporte específico
        
        Args:
            report_id: ID del reporte
            movements: Lista de movimientos actualizados
            
        Returns:
            bool: True si se actualizó correctamente
        """
        try:
            # Crear backup antes de actualizar
            if not self.backup_file():
                print("Advertencia: No se pudo crear backup antes de actualizar movimientos")
            
            workbook = openpyxl.load_workbook(self.file_path)
            
            # Eliminar movimientos existentes del reporte
            if "ingresos_retiros" in self.sheets and self.sheets["ingresos_retiros"] in workbook.sheetnames:
                movimientos_sheet = workbook[self.sheets["ingresos_retiros"]]
                rows_to_delete = []
                
                for row_idx, row in enumerate(movimientos_sheet.iter_rows(min_row=2), start=2):
                    if row[0].value == report_id:  # Columna A contiene ID_Reporte
                        rows_to_delete.append(row_idx)
                
                # Eliminar filas en orden inverso para no afectar índices
                for row_idx in sorted(rows_to_delete, reverse=True):
                    movimientos_sheet.delete_rows(row_idx, 1)
                
                # Agregar los nuevos movimientos
                for movement in movements:
                    # Manejar tanto objetos Pydantic como diccionarios
                    if hasattr(movement, 'dict'):
                        movement_dict = movement.dict()
                    else:
                        movement_dict = movement
                    
                    new_row = [
                        report_id,
                        movement_dict.get('nombre_empleado', ''),
                        movement_dict.get('cargo', ''),
                        movement_dict.get('estado', '')
                    ]
                    movimientos_sheet.append(new_row)
                
                print(f"Actualizados {len(movements)} movimientos para reporte {report_id}")
            
            # Guardar cambios
            workbook.save(self.file_path)
            workbook.close()
            
            return True
            
        except Exception as e:
            print(f"Error actualizando movimientos del reporte {report_id}: {e}")
            try:
                workbook.close()
            except:
                pass
            return False


# Instancia global del manejador
excel_handler = ExcelHandler()